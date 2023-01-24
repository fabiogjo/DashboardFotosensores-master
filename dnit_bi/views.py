from cgitb import handler
from decimal import ROUND_DOWN, Decimal
import json
from datetime import date, datetime, timedelta
import os
from django.shortcuts import render, redirect, get_object_or_404
from requests.auth import HTTPBasicAuth
import requests as requests
from .models import Paralisado, Offline, Equipamento, Ticket_freshdesk, Indice_desempenho, Estudo_tecnico
from usuarios.models import Usuario
from django.db.models.functions import Now
from django.http import JsonResponse, HttpResponseRedirect
from django.db.models import Count, F, IntegerField, DateField, Min, DateTimeField, Func, Avg, DecimalField, Sum, Max, Min
from .forms import *
from django.utils import timezone
from django.utils.translation import gettext
from django.contrib.auth.decorators import login_required
import io
import base64
import PIL.Image as Image
import pyimgur
from bs4 import BeautifulSoup
import openpyxl
from django.contrib.postgres.search import SearchVector


# PAGINAS
def index(request):
    return render(request, 'index.html')


def paralisados(request):
    paralisados = Paralisado.objects.exclude(status="Deferida (Encerramento)").order_by('data_abertura')
    form = Edita_paralisado_form()

    dados = {

        'paralisados': paralisados,
        'form': form

    }

    return render(request, 'paralisados.html', dados)


def offlines(request):
    offlines = Offline.objects.all().order_by('offline_desde')

    dados = {

        'offlines': offlines

    }

    return render(request, 'offlines.html', dados)

@login_required(login_url="/usuarios/login")
def tickets_freshdesk(request):
    
    if request.user.is_authenticated:    

        equipamentos = Equipamento.objects.exclude(situacao_faixa="Em Instalação").annotate(Count('ticket_freshdesk'), Max('ticket_freshdesk__dias_aberto')).order_by('ticket_freshdesk__dias_aberto__max')                     

        form_ticket = Cria_ticket_form
        
        estudos_tecnicos = Estudo_tecnico.objects.exclude(situacao="Aprovado").exclude(situacao="Anulado").order_by('municipio__lote')

        
        
        dados = {
            
            'equipamentos': equipamentos,
            'form': form_ticket,
            'estudos_tecnicos':estudos_tecnicos,

        }

        return render(request, 'tickets_freshdesk_new.html', dados)
    
    else:
        
        return render(request, 'login')


def indice_desempenho(request):
    return render(request, 'indice_desempenho.html')


def detalhar_equipamento(request, equipamento_id):
    if request.method == "POST":

        id = request.POST.get('id')
        anotacao = request.POST.get('anotacao')
        equipamento = request.POST.get('equipamento_id')
        status = request.POST.get('status_ticket')
        user_key = request.POST.get('user_key')
        
        equipamento = Equipamento.objects.get(id=equipamento_id)

        base_url = "https://fotosensores-dnit.freshdesk.com/api/v2/tickets"

        headers = {'Content-Type': 'application/json'}

        auth = HTTPBasicAuth(user_key, '')

        ticket = Ticket_freshdesk.objects.filter(id_ticket=id).values()

        ticket_status = ticket[0]['status']

        status_dict = {
            'Aberto': 2,
            'Pendente': 3,
            'Resolvido': 4,
            'Fechado': 5,
        }
         
        
        soup = BeautifulSoup(anotacao)
        for img in soup.findAll('img'):
            img['src'] = base64_to_file(img['src'])
        anotacao =  str(soup)
        

        if status == ticket_status:

            anotacao = {

                'body': f'{anotacao}'

            }

            r = requests.post(f'{base_url}/' + str(id) + '/reply', auth=auth,
                          headers=headers, data=json.dumps(anotacao))
            
            print(r)      
            
            nova_notificacao = Notificacao(usuario = request.user, equipamento=equipamento, acao=f'adicionou uma anotacao ao ticket {id}')
            
            nova_notificacao.save()
            
        else:

            anotacao = {

                'body': f'{anotacao}'

            }

            requests.post(f'{base_url}/' + str(id) + '/reply', auth=auth,
                          headers=headers, data=json.dumps(anotacao))

            ticket = {

                'status': status_dict[status],
            }

            requests.put(f'{base_url}/' + str(id), auth=auth,
                         headers=headers, data=json.dumps(ticket))

            ticket = Ticket_freshdesk.objects.get(id_ticket=id)

            ticket.status = status

            ticket.save()
            
            nova_notificacao = Notificacao(usuario = request.user, equipamento=equipamento, acao=f'adicionou uma anotacao ao ticket {id} e status como {status}')
            
            nova_notificacao.save()


    equipamento = get_object_or_404(Equipamento, pk=equipamento_id)
    
            
    equipamento.fotos = []  
            
    faixas = equipamento.get_faixas()
        
    if len(faixas) > 1:
        
        for faixa in faixas:
                        
            path = f'dashboard/static/equips/{equipamento.codigo}/P-{faixa.sentido}-{faixa.numero}/'
                        
            for (root,dirs,files) in os.walk(path, topdown=True):
                        
                for file in files:
                            
                    if file == 'Thumbs.db':
                        continue
                        
                    equipamento.fotos.append(f'{equipamento.codigo}/P-{faixa.sentido}-{faixa.numero}/{file}')
                    
    else:
                
        path = f'dashboard/static/equips/{equipamento.codigo}/'
                    
        for (root,dirs,files) in os.walk(path, topdown=True):
                        
            for file in files:
                        
                if file == 'Thumbs.db':
                    continue
                                
                equipamento.fotos.append(f'{equipamento.codigo}/{file}')

    
    equipamento_a_exibir = {
        
        'equipamento': equipamento,
        
    }
    
    return render(request, 'equipamento.html', equipamento_a_exibir)


def busca(request):
    
    lista_de_equipamentos = Equipamento.objects.all()
    
    nome_a_buscar = ''
    
    serial_a_buscar = ''

    municipio_a_buscar = ''
    
    form_ticket = Cria_ticket_form

    
    if 'buscaGeral' in request.GET:
        
        a_buscar = request.GET['buscaGeral'].strip()
        
        lista_de_equipamentos = Equipamento.objects.filter(Q(municipio__nome__icontains=a_buscar) | Q(numero_de_serie__icontains=a_buscar) | Q(codigo__icontains=a_buscar) | Q(ticket_freshdesk__id_ticket__icontains=a_buscar) | Q(ticket_freshdesk__agente__icontains=a_buscar)).distinct().annotate(Count('ticket_freshdesk'), Max('ticket_freshdesk__dias_aberto')).order_by('ticket_freshdesk__dias_aberto__max')    
        

        
        dados = {
            
            'equipamentos' : lista_de_equipamentos,
            'form': form_ticket,
        }
        
        print(lista_de_equipamentos)

        return render(request, 'busca.html', dados)
        

    
    if set(['buscaAgente', 'buscaSerial', 'buscaMunicipio', 'buscaLote']).issubset( request.GET ):
        
        nome_a_buscar = request.GET['buscaAgente']
        
        serial_a_buscar = request.GET['buscaSerial']
        
        municipio_a_buscar = request.GET['buscaMunicipio']
        
        lote_a_buscar = request.GET['buscaLote']
        
        if nome_a_buscar == '' and serial_a_buscar == '' and municipio_a_buscar == '' and lote_a_buscar == '':
            
            #lista_de_equipamentos = lista_de_equipamentos.exclude(situacao_faixa="Em Instalação").annotate(Count('ticket_freshdesk'), Max('ticket_freshdesk__dias_aberto')).order_by('ticket_freshdesk__dias_aberto__max')  

            return redirect('tickets_freshdesk')
            
        else:
        
            a_pesquisar = []
            
            a_pesquisar.append(nome_a_buscar)
            
            a_pesquisar.append(municipio_a_buscar)
            
            a_pesquisar.append(serial_a_buscar)
            
            a_pesquisar.append(lote_a_buscar)
            
            rodada = 0
            
            for item in a_pesquisar:
                
                rodada += 1
                                
                if rodada == 1 and item != '':
            
                    lista_de_equipamentos = lista_de_equipamentos.filter(ticket_freshdesk__agente__icontains=item).distinct().annotate(Count('ticket_freshdesk'), Max('ticket_freshdesk__dias_aberto')).order_by('ticket_freshdesk__dias_aberto__max')
                    
                    continue
                    
                if rodada == 2 and item != '':
                    
                    lista_de_equipamentos = lista_de_equipamentos.filter(municipio__nome__icontains=item).distinct().annotate(Count('ticket_freshdesk'), Max('ticket_freshdesk__dias_aberto')).order_by('ticket_freshdesk__dias_aberto__max')
                    
                    continue
                    
                if rodada == 3 and item != '':
                    
                    lista_de_equipamentos = lista_de_equipamentos.filter(numero_de_serie=item).distinct().annotate(Count('ticket_freshdesk'), Max('ticket_freshdesk__dias_aberto')).order_by('ticket_freshdesk__dias_aberto__max')
                    
                    continue
                
                if rodada == 4 and item != '':
                    
                    lista_de_equipamentos = lista_de_equipamentos.filter(municipio__lote__numero__=item).distinct().annotate(Count('ticket_freshdesk'), Max('ticket_freshdesk__dias_aberto')).order_by('ticket_freshdesk__dias_aberto__max')
                    
                    break
        
    
                  
    dados = {
        
        
        
        'equipamentos' : lista_de_equipamentos,
        
        
    }
    
    print(lista_de_equipamentos)

    return render(request, 'tickets_freshdesk_new.html', dados)


def upload_id(request):
    if request.method == 'POST':      
        
        Indice_desempenho.objects.all().delete()
        
        faixas = Faixa.objects.all()
        
        for f in request.FILES.getlist("uploadedFile"):
            
            dict_indice_desempenho = {}
            
            wb = openpyxl.load_workbook(f)
            
            worksheet = wb["Sheet1"]
        
            for faixa in faixas:
            
                for linha in worksheet.iter_rows(min_row=4, min_col=1, max_col=60, values_only=True):
                    
                    if faixa.equipamento.codigo not in dict_indice_desempenho:
                        dict_indice_desempenho[faixa.equipamento.codigo] = {}
                    
                    if faixa.equipamento.codigo != linha[6] and f'P-{faixa.sentido}-{faixa.numero}' != linha[8]:
                        
                        continue    
                                
                    elif faixa.equipamento.codigo == linha[6] and f'P-{faixa.sentido}-{faixa.numero}' == linha[8]: 
                                
                        dict_indice_desempenho[faixa.equipamento.codigo][linha[8]] = {
                                    'equipamento': linha[6],
                                    'remessas_infracoes': linha[11],
                                    'testes': linha[12],
                                    'infracoes': linha[13],
                                    'situacao': linha[14],
                                    'validas': linha[15],
                                    'invalidas': linha[16],
                                    'remessas_passagens': linha[17],
                                    'icid': float(round(linha[24], 2)),
                                    'icin': float(round(linha[28], 2)),
                                    'ievri': float(round(linha[34], 2)),
                                    'ievdt': float(round(linha[42], 2)),
                                    'ilpd': float(round(linha[45], 2)),
                                    'ilpn': float(round(linha[48], 2)),
                                    'icv': float(round(linha[51], 2)),
                                    'ief': float(round(linha[52], 2)),
                                    'periodo': (linha[53]).date(),
                                    'dias': linha[54],
                                    'nht': linha[55],
                                    'nho': linha[56],
                                    'idf': float(round(linha[57], 2)),
                                    'indice_desempenho': float(round(linha[58], 3)),
                                    'atualizado': datetime.now(),    
                                    
                                        
                                    'equipamento_id':faixa.equipamento.id,
                                    'faixa_id':faixa.id

                            }
            
            for equipamento in dict_indice_desempenho:
                
                for faixa in dict_indice_desempenho[equipamento]:            
                    
                    r = dict_indice_desempenho[equipamento][faixa]
            
                    indice = Indice_desempenho(faixa_id=r['faixa_id'], remessas_infracoes=r['remessas_infracoes'], testes=r['testes'], infracoes=r['infracoes'], situacao=r['situacao'], validas=r['validas'], invalidas=r['invalidas'], remessas_passagens=r['remessas_passagens'], icid=r['icid'], icin=r['icin'], ievri=r['ievri'], ievdt=r['ievdt'], ilpd=r['ilpd'], ilpn=r['ilpn'], icv=r['icv'], ief=r['ief'], periodo=r['periodo'], dias=r['dias'], nht=r['nht'], nho=r['nho'], idf=r['idf'], indice_desempenho=r['indice_desempenho'], atualizado=r['atualizado'])
                    indice.save()
                    
        return redirect('tickets_freshdesk')
        
    return render(request, 'upload_id.html')


def upload_et(request):
    
    if request.method == 'POST':      
        
        file = request.FILES.get("uploadedFile")
        
        Estudo_tecnico.objects.all().delete()
        
        equipamentos = Equipamento.objects.all()  
            
        dict_et = {}      
        
        wb = openpyxl.load_workbook(file)
            
        worksheet = wb["Sheet1"]
                
        for linha in worksheet.iter_rows(min_row=2, min_col=1, max_col=9, values_only=True):
            
            municipio = Municipio.objects.filter(nome=linha[5])
            
            if not municipio:
                
                numero_lote_planilha = linha[1][0:1]
                
                print(numero_lote_planilha)
                
                
                novo_municipio = Municipio(nome=linha[5])
                
                novo_municipio.save()
                             
                municipio = Municipio.objects.filter(nome=linha[5])[0]
                    
            
            if municipio:
            
                result = f'{linha[4]:.3f}'
                          
                dict_et[result] = {
                            
                    'municipio': municipio[0].id,
                    
                    'km': result,
                    
                    'br': linha[3],
                    
                    'tipo_equipamento': linha[7],
                            
                    'codigo': linha[0],
                            
                    'situacao': linha[8],
                    
                    'faixas': linha[6],
                                          
                    'ultima_atualizacao_situacao': None,
                            
                    'atualizado': datetime.now(),

                }
            
        for i in dict_et:    
            
            r = dict_et[i]
                
            et = Estudo_tecnico(faixas=r['faixas'], km=r['km'],br=r['br'], tipo_equipamento=r['tipo_equipamento'], codigo=r['codigo'], situacao=r['situacao'], ultima_atualizacao_situacao=r['ultima_atualizacao_situacao'], atualizado=r['atualizado'], municipio_id=r['municipio'])
                   
            et.save()
            
        return redirect('tickets_freshdesk')
                
        
    return render(request, 'upload_et.html')


def upload_paralisados(request):
    if request.method == 'POST':      
        
        file = request.FILES.get("uploadedFile")
        
        wb = openpyxl.load_workbook(file)
            
        worksheet = wb["Sheet1"]
        
                
        for linha in worksheet.iter_rows(min_row=3, min_col=1, max_col=13, values_only=True):
            
            codigo_equipamento_planilha = linha[5]
            
            data_de_abertura_paralisacao_planilha = linha[10]
            
            motivo_paralisacao_planilha = linha[11]
            
            sentido_faixa_planilha = linha[4]
            
            
            equipamento = Equipamento.objects.filter(codigo=codigo_equipamento_planilha).first()
            
            if equipamento:
                faixas = equipamento.get_faixas()
            
            else:
                print(f'equipamento {codigo_equipamento_planilha} não consta na base de dados')
                continue
            
            for faixa in faixas:
                
                keyPlanilha = f'{codigo_equipamento_planilha} - {sentido_faixa_planilha}'
                
                keyBanco = f'{equipamento.codigo} - P-{faixa.sentido}-{faixa.numero}'
                
                if keyBanco != keyPlanilha:
                    continue
                    
                paralisado = Paralisado.objects.filter(faixa_id=faixa.id).exclude(status="Deferida (Encerramento)").first()
                
                if paralisado and paralisado.faixa.equipamento.numero_de_serie == 11457:
                    print(paralisado.faixa.equipamento.get_paralisacao_status())
                
                if paralisado and data_de_abertura_paralisacao_planilha is None and paralisado.faixa_id == faixa.id:
                        
                    paralisado.data_encerramento = datetime.today()
                        
                    paralisado.status = "Deferida (Encerramento)"
                        
                    paralisado.save()
                    
                    print(f'Fechada paralisação {paralisado}')
                    
                    continue
                
                    
                elif paralisado and data_de_abertura_paralisacao_planilha != None:
                    continue
                
                
                elif paralisado is None and data_de_abertura_paralisacao_planilha != None:
                    
                    novo_paralisado = Paralisado(faixa_id=faixa.id, data_abertura=data_de_abertura_paralisacao_planilha, motivo=motivo_paralisacao_planilha, status="Deferida (Ínicio)", situacao="Em ação")
                    
                    novo_paralisado.save()
                    
                    print(f'Aberta paralisação {novo_paralisado}')
                    
                    continue
            
        return redirect('tickets_freshdesk')
                
        
    return render(request, 'upload_paralisado.html')
    
  
def profile(request):
    
    user = request.user
    
    user_full_name = f'{user.nome} {user.sobrenome}'
    
    numero_de_tickets = Ticket_freshdesk.objects.filter(agente=user_full_name)
    
    user_tickets = Equipamento.get_user_tickets(user_full_name)
    
    equipamentos = Equipamento.objects.filter(ticket_freshdesk__agente=user_full_name).distinct().annotate(Count('ticket_freshdesk'), Max('ticket_freshdesk__dias_aberto')).order_by('ticket_freshdesk__dias_aberto__max')
    
    form_ticket = Cria_ticket_form
    
    dados = {
        
        'numero_de_tickets':numero_de_tickets.count(),
        
        'user_tickets':user_tickets,
        
        'equipamentos':equipamentos,
        
        'user_full_name': user_full_name,
        
        'form': form_ticket
        
    }
    
    return render(request, 'profile.html', dados)  
   
    
def testes(request):
    
    if request.method == "POST":

        id = request.POST.get('id')
        anotacao = request.POST.get('anotacao')
        equipamento = request.POST.get('equipamento_id')
        status = request.POST.get('status_ticket')
        user_key = request.POST.get('user_key')
        
        equipamento = Equipamento.objects.get(id=1809)

        base_url = "https://fotosensores-dnit.freshdesk.com/api/v2/tickets"

        headers = {'Content-Type': 'application/json'}

        auth = HTTPBasicAuth(user_key, '')

        ticket = Ticket_freshdesk.objects.filter(id_ticket=id).values()

        ticket_status = ticket[0]['status']

        status_dict = {
            'Aberto': 2,
            'Pendente': 3,
            'Resolvido': 4,
            'Fechado': 5,
        }
         
        
        soup = BeautifulSoup(anotacao)
        for img in soup.findAll('img'):
            img['src'] = base64_to_file(img['src'])
        anotacao =  str(soup)
        

        if status == ticket_status:

            anotacao = {

                'body': f'{anotacao}'

            }

            r = requests.post(f'{base_url}/' + str(id) + '/reply', auth=auth,
                          headers=headers, data=json.dumps(anotacao))
            
            print(r)      
            
            nova_notificacao = Notificacao(usuario = request.user, equipamento=equipamento, acao=f'adicionou uma anotacao ao ticket {id}')
            
            nova_notificacao.save()
            
        else:

            anotacao = {

                'body': f'{anotacao}'

            }

            requests.post(f'{base_url}/' + str(id) + '/reply', auth=auth,
                          headers=headers, data=json.dumps(anotacao))

            ticket = {

                'status': status_dict[status],
            }

            requests.put(f'{base_url}/' + str(id), auth=auth,
                         headers=headers, data=json.dumps(ticket))

            ticket = Ticket_freshdesk.objects.get(id_ticket=id)

            ticket.status = status

            ticket.save()
            
            nova_notificacao = Notificacao(usuario = request.user, equipamento=equipamento, acao=f'adicionou uma anotacao ao ticket {id} e status como {status}')
            
            nova_notificacao.save()


    equipamento = get_object_or_404(Equipamento, pk=1809)
    
            
    equipamento.fotos = []  
            
    faixas = equipamento.get_faixas()
        
    if len(faixas) > 1:
        
        for faixa in faixas:
                        
            path = f'dashboard/static/equips/{equipamento.codigo}/P-{faixa.sentido}-{faixa.numero}/'
                        
            for (root,dirs,files) in os.walk(path, topdown=True):
                        
                for file in files:
                            
                    if file == 'Thumbs.db':
                        continue
                        
                    equipamento.fotos.append(f'{equipamento.codigo}/P-{faixa.sentido}-{faixa.numero}/{file}')
                    
    else:
                
        path = f'dashboard/static/equips/{equipamento.codigo}/'
                    
        for (root,dirs,files) in os.walk(path, topdown=True):
                        
            for file in files:
                        
                if file == 'Thumbs.db':
                    continue
                                
                equipamento.fotos.append(f'{equipamento.codigo}/{file}')

    
    equipamento_a_exibir = {
        
        'equipamento': equipamento,
        
    }
    
    return render(request, 'partials/testes.html', equipamento_a_exibir)

    
# FUNCTIONS

def edita_paralisados(request):
    if request.method == "POST":
        id = request.POST.get('id')
        data_encerramento = request.POST.get('data_encerramento')
        status = request.POST.get('status')
        paralisado = Paralisado.objects.get(id=id)

        if len(data_encerramento) == 0:
            data_encerramento = None

        paralisado.status = status

        paralisado.data_encerramento = data_encerramento

        paralisado.save()

        return redirect('paralisados')


def altera_situacao_paralisado(request):
    if request.method == "POST":
        
        id = request.POST.get('equipId')
        
        situacao = request.POST.get('situacao')
        
        print(situacao)
        
        equipamento = Equipamento.objects.get(id=id)
        
        user = Usuario.objects.get(id=request.user.id)
        
        faixas = Faixa.objects.filter(equipamento_id=id)
             
        for faixa in faixas:
            
            paralisado = Paralisado.objects.filter(faixa_id=faixa.id).exclude(status="Deferida (Encerramento)").first()
            
            if not paralisado:
                continue
            
            nova_notificacao = Notificacao(usuario=user, equipamento=equipamento, acao=f'alterou a situação da paralisação de {paralisado.situacao} para {situacao}')

            nova_notificacao.save()
            
            paralisado.situacao = situacao
            
            paralisado.save()
            

        return redirect('tickets_freshdesk')


def cria_paralisado(request):
    if request.method == 'POST':
        form = Edita_paralisado_form(request.POST)
        if form.is_valid():
            paralisado = form.save(commit=False)
            paralisado.data_encerramento = None
            paralisado.save()
            return redirect('paralisados')


def base64_to_file(byte_data):
    
    start = byte_data.find('base64,') + 7
        
    recorte_start = byte_data[start:]
        
    end = recorte_start.find('"')
        
    recorte_end = recorte_start[:end]
    
    byte_data = recorte_end

    
    b = base64.b64decode(byte_data + "==")
    
    img = Image.open(io.BytesIO(b))
    
    img.save('to_upload_img.png')
    
    CLIENT_ID = "a8b9b4f494c28b5"
    
    PATH = 'to_upload_img.png'
    
    im = pyimgur.Imgur(CLIENT_ID)
    
    uploaded_image = im.upload_image(PATH, title="Uploaded with PyImgur")
    
    return uploaded_image.link
    
    
def cria_ticket(request):
    if request.method == 'POST':

        form = Cria_ticket_form(request.POST)     

        if form.is_valid():

            tickets_freshdesk = form.save(commit=False)
            
            base_url = "https://fotosensores-dnit.freshdesk.com/api/v2/tickets"

            headers = {'Content-Type': 'application/json'}
            
            user_key = request.POST['user_key']
            
            user_email = request.POST['user_email']

            auth = HTTPBasicAuth(user_key, '')

            hoje = datetime.today()

            contato = user_email

            prioridade_dict = {

                'Baixa': 1,
                'Media': 2,
                'Alta': 3,
                'Urgente': 4,

            }


            assunto = f'{tickets_freshdesk.equipamento.numero_de_serie} - {tickets_freshdesk.equipamento.codigo} - BR {tickets_freshdesk.equipamento.br} - KM {tickets_freshdesk.equipamento.km} - {tickets_freshdesk.equipamento.municipio} - {tickets_freshdesk.assunto}'


            equipamento = Equipamento.objects.get(id=tickets_freshdesk.equipamento_id)
        
            nome_responsavel = ''
            
            infraestrutura_tipos = ["Equipamento sem energia", "Equipamento sem energia", "Instalação / Reparo de energia eletrica", "Instalação / Reparo de energia eletrica","Poda / Roçada", "Poda / Roçada","Instalação / Reparo de cabo lógico", "Instalação / Reparo de cabo lógico","Implantação / ajuste de sinalização", "Implantação / ajuste de sinalização","infraestrutura", "infraestrutura","Implantação/Reparo de sinalização vertical", "Implantação/Reparo de sinalização vertical","PISTA DANIFICADA SEM CONDIÇÕES PRA REFAZER LAÇOS", "PISTA DANIFICADA SEM CONDIÇÕES PRA REFAZER LAÇOS","Implantação/Reparo de sinalização horizontal", "Implantação/Reparo de sinalização horizontal"]
                
            
            if tickets_freshdesk.tipo in infraestrutura_tipos:
                
                if equipamento.get_estado() == 'RS':
                    
                    id_responsavel = 67025985129
                    grupo = 67000228873                 
                    nome_responsavel = f'Adilson Rodrigues'
                    
                elif equipamento.get_estado() == 'SC':
                    
                    nome_responsavel = ''
                    id_responsavel = None
                    grupo = 67000228873
                    
            elif tickets_freshdesk.tipo == 'Service Task':
                
                id_responsavel = int(equipamento.municipio.setor.responsavel.id_freshdesk)
            
                nome_responsavel = f'{equipamento.municipio.setor.responsavel.nome} {equipamento.municipio.setor.responsavel.sobrenome}'
                
                grupo = equipamento.municipio.setor.responsavel.grupo_id
                    
            else:
                grupo = 67000215926
                id_responsavel = 67033345228
                nome_responsavel = f'Monitoramento Operacional'

            
            if grupo != None:
                grupo = int(grupo)
            
            local_servico = f'{equipamento.latitude},{equipamento.longitude}'
            

            td = timedelta(hours=8)

            hoje = datetime.now() + td

            hoje = hoje.isoformat()
            
            
            soup = BeautifulSoup(tickets_freshdesk.descricao)
            for img in soup.findAll('img'):
                img['src'] = base64_to_file(img['src'])
            descricao =  str(soup)

            
            if tickets_freshdesk.tipo != 'Service Task':
                exemple = {
                    'subject': assunto,
                    'description': descricao,
                    'email': contato,
                    'priority': prioridade_dict[tickets_freshdesk.prioridade],
                    'status': 2,                   
                    'group_id': grupo,
                    'responder_id': id_responsavel,
                    'type': tickets_freshdesk.tipo
                }
                
            else:
                exemple = {
                    'subject': assunto,
                    'description': descricao,
                    'email': contato,
                    'priority': prioridade_dict[tickets_freshdesk.prioridade],
                    'status': 2,                   
                    'group_id': grupo,
                    'responder_id': id_responsavel,
                    'type': 'Service Task',
                    "custom_fields": {
                        "cf_fsm_contact_name": contato,
                        "cf_fsm_phone_number": "54996794368",
                        "cf_fsm_service_location": local_servico,
                        "cf_fsm_appointment_start_time": hoje,
                        "cf_fsm_appointment_end_time": hoje
                    },  
                }

            exemplo = json.dumps(exemple)

            response = requests.post(f'{base_url}', auth=auth,
                                     headers=headers, data=exemplo)
            

            id_ticket_criado = int(response.headers.get('Location')[-5:])

            data_formatada = datetime.now().strftime('%Y-%m-%d')
            
            
            tickets_freshdesk.data_criacao = data_formatada
            tickets_freshdesk.agente = nome_responsavel
            tickets_freshdesk.dias_aberto = 1
            tickets_freshdesk.prioridade = tickets_freshdesk.prioridade
            tickets_freshdesk.status = "Aberto"
            tickets_freshdesk.id_ticket = id_ticket_criado
            tickets_freshdesk.save()
            
            nova_notificacao = Notificacao(usuario = request.user, equipamento=equipamento, acao=f'criou ticket {id_ticket_criado} - {tickets_freshdesk.assunto} para o agente {tickets_freshdesk.agente}')
            
            nova_notificacao.save()

        return redirect('tickets_freshdesk')


def close_ticket(ticket_id):
    base_url = "https://fotosensores-dnit.freshdesk.com/api/v2/tickets"

    headers = {'Content-Type': 'application/json'}

    auth = HTTPBasicAuth('G8jg3T9KHgD5GlyaQmVq', '')

    ticket = {

        'status': 4,
    }

    response = requests.put(f'{base_url}/' + str(ticket_id), auth=auth,
                            headers=headers, data=json.dumps(ticket))

    return response


def adiciona_caminho_fotos_aos_equipamentos(equipamentos):
    
    for equipamento in equipamentos:
            
            equipamento.fotos = []  
                
            faixas = equipamento.get_faixas()

                
            if len(faixas) > 1:
                    
                for faixa in faixas:
                
                        
                    path = f'dashboard/static/equips/{equipamento.codigo}/P-{faixa.sentido}-{faixa.numero}/'
                        
                    for (root,dirs,files) in os.walk(path, topdown=True):
                            
                        for file in files:
                                
                            if file == 'Thumbs.db':
                                continue
                            
                            equipamento.fotos.append(f'{equipamento.codigo}/P-{faixa.sentido}-{faixa.numero}/{file}')
                            
                            break

            else:
                    
                path = f'dashboard/static/equips/{equipamento.codigo}/'
                        
                for (root,dirs,files) in os.walk(path, topdown=True):
                            
                    for file in files:
                            
                        if file == 'Thumbs.db':
                            continue
                                    
                        equipamento.fotos.append(f'{equipamento.codigo}/{file}')
                        
                        break


def upload_agentes(request):
    base_url = "https://fotosensores-dnit.freshdesk.com/api/v2/agents?per_page=100"

    headers = {'Accept': 'application/json'}

    auth = HTTPBasicAuth('G8jg3T9KHgD5GlyaQmVq', '')

    response = requests.get(base_url, auth=auth, headers=headers)

    agentes = response.json()
    
    for agente in agentes:
        
        
        s = Usuario.objects.filter(nome=agente['contact']['name'])
        
        if agente['type'] == 'field_agent':
            
            q = Agente.objects.filter(id_freshdesk=agente['id'])
            
            if not q:
                
                new_agent = Agente(id_freshdesk=agente['id'], nome=agente['contact']['name'], sobrenome='', email=agente['contact']['email'], grupo='Técnicos', grupo_id=67000573741, celular=0)
                
                new_agent.save()
             
            
                
                
    
    return render(request, 'partials/testes.html')


def nova_anotacao_equipamento(request, equipamento_id, user_id):
    
     if request.method == "POST":
         
        anotacao = request.POST.get('anotacao')
        
        anotacao = anotacao.replace('<p>', '')
        
        anotacao= anotacao.replace('</p>', '')
        
        anotacao= anotacao.replace('&nbsp;', '')
        
        user_id = request.user.id
        
        user = Usuario.objects.get(id=user_id)
        
        soup = BeautifulSoup(anotacao)
        
        if soup.img:
            anotacao = soup.img.decompose()      
            anotacao = str(soup)
        
        
        equipamento = Equipamento.objects.get(id=equipamento_id)
        
        usuario = Usuario.objects.get(id=user_id)
        
        nova_anotacao = EquipamentoAnotacoes(equipamento=equipamento, usuario=usuario, anotacao=anotacao)
        
        nova_anotacao.save()
        
        noticacao = Notificacao(usuario=user, equipamento=equipamento, acao="adicionou uma anotação")
        
        noticacao.save()
        
        return redirect('detalhar_equipamento', equipamento.id)
    
# JSON
def json_indice_desempenho(request):
    indices = Indice_desempenho.objects.all()
    data = [indice.get_data() for indice in indices]
    response = {'data': data}
    return JsonResponse(response)
    

def json_notificacoes(request):
    notificacoes = Notificacao.objects.all()
    data = [notificacao.get_data() for notificacao in notificacoes]
    response = {'data': data}
    return JsonResponse(response)


def json_paralisados(request):
    paralisados = Paralisado.objects.all().exclude(status='Deferida (Encerramento)')
    data = [paralisado.get_data() for paralisado in paralisados]
    response = {'data': data}
    return JsonResponse(response)


def json_equipamentos(request):
    equipamentos = Equipamento.objects.all()
    data = [equipamento.get_data() for equipamento in equipamentos]
    response = {'data': data}
    return JsonResponse(response)


def tickets_por_tipo(request):
    tipos = Ticket_freshdesk.objects.values('tipo').annotate(Count('tipo')).order_by('-tipo__count')
    return JsonResponse(list(tipos), safe=False)


def tickets_por_setor(request):
    agentes = Ticket_freshdesk.objects.values('agente').annotate(Count('id_ticket')).order_by('-id_ticket__count')
    return JsonResponse(list(agentes), safe=False)


def tickets_por_lote(request):
    lote = Equipamento.objects.values('lote').annotate(Count('ticket_freshdesk')).order_by('-ticket_freshdesk__count')
    return JsonResponse(list(lote), safe=False)


def paralisacoes_por_mes(request):
    # lote = Paralisado.objects.annotate(month=ExtractMonth('data_abertura'))\
    #     .values('month')\
    #     .annotate(c=Count('id'))\
    #     .order_by()
    d = []

    # arbitrary starting dates
    year = 2019
    month = 12

    cyear = datetime.now().year
    cmonth = datetime.now().month

    while year <= cyear:
        while (year < cyear and month <= 12) or (year == cyear and month <= cmonth):
            paralisacoes = Paralisado.objects.filter(data_abertura__year=year, data_abertura__month=month).aggregate(
                Count('id'))
            desparalisacoes = Paralisado.objects.filter(data_encerramento__year=year,
                                                        data_encerramento__month=month).aggregate(Count('id'))
            d.append({
                'year': year,
                'month': month,
                'paralisados': paralisacoes['id__count'] or 0,
                'desparalisados': desparalisacoes['id__count'] or 0,
            })
            month += 1
        month = 1
        year += 1

    return JsonResponse(list(d), safe=False)


def indice_por_tecnico(request):
    class Round(Func):
        function = 'ROUND'
        template = '%(function)s(%(expressions)s, 2)'

    r = Equipamento.objects \
        .values('municipio__setor__responsavel__nome', 'tipo_equipamento') \
        .annotate(media=(Round(Avg('faixa__indice_desempenho__indice_desempenho')))) \
        .order_by('-media')

    return JsonResponse(list(r), safe=False)


def consulta_agente(request, id_agente):

    r = Agente.objects.filter(id_freshdesk=id_agente).values('nome', 'sobrenome', 'id_freshdesk')

    return JsonResponse(list(r), safe=False)


def json_faixas_por_setor(request):
    faixas = Equipamento.objects.values('municipio__setor').annotate(Count('numero_de_serie')).order_by('-numero_de_serie__count')
    return JsonResponse(list(faixas), safe=False)


def get_notificacoes(request):
    # request should be ajax and method should be GET.
    if request.is_ajax and request.method == "GET":
        # get the nick name from the client side.
        # check for the nick name in the database.
        notificacoes = Notificacao.objects.all().order_by('-created_at').values('usuario', 'equipamento', 'acao')
        # if nick_name found return not valid new friend
        return JsonResponse(list(notificacoes), safe=False)

    return JsonResponse({}, status = 400)